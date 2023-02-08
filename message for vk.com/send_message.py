try:
    import sys
    # Импортируем наш интерфейс
    from GUI import *
    from PyQt5 import QtCore, QtGui, QtWidgets
    from PyQt5.QtGui import QColor
    from python_rucaptcha import ImageCaptcha
    from pika import exceptions
    from python_rucaptcha import *
    import threading
    import schedule
    import time
    import json
    import vk_api
    from PyQt5.QtCore import QThread
    import random
    from datetime import datetime
    from openpyxl import load_workbook
    from auth_vk_browser import auth_vk_driver
    from selenium.webdriver import Chrome
    from selenium.webdriver.chrome.options import Options
    from selenium.webdriver.support.wait import WebDriverWait
    import random
    from generate_message import gen_text
    import urllib
except Exception as err:
    input(f'err: {err}')

#tt : QtWidgets.QTextEdit
API_KEY = '26d7985260270f8ea8b4ab36de2839ac'
FILE_SAVE_ALL = 'saveData.json'
PAUSE = 5



print("Random number with system time")
random.seed(datetime.now())
FILE_BASE_EXEL='base.xlsx'
FILE_BAD = 'bad' + str(datetime.now().strftime("%d-%m-%Y %H_%M")) + '.txt'
from datetime import datetime


def getIMG(text:str):
    result = ''
    mas = text.split('\n')
    for s in mas:
        if s.find('photo') != -1:
            #нашли ссылку изобр. из albums
            photo = s[s.find('photo'):]
            if photo.find('%')!=-1:
                photo = photo[:photo.find('%')]
            result = result +  photo +','
            pass
            continue
        addLog(f'Ошибка получения изображения ссылка {s}')

    return result


#print(getIMG('https://vk.com/albums-70063346?z=photo-70063346_459688665%2Fphotos-70063346\nhttps://vk.com/albums-70063346?z=photo-70063346_459688665%2Fphotos-70063346'))

#input('stop')


def saveBad(s):
    with open(FILE_BAD, 'a', encoding='utf-8') as f:
        f.write(str(s)+'\n')


data_auth = {
    'login':'',
    'password':'',
    'token':''
}

PAUSE_MIN=1
PAUSE_MAX=5

FILE_BASE = 'base.txt'
vk = None
memoLog = None
buttonAuth = None
imageCaptcha = None
buttonCaptcha = None
editCaptcha = None
editPassword = None
editLogin = None
listWidget = None
buttonStart = None
buttonPause = None
buttonLoadAccs = None
memoText = QtWidgets.QTextEdit
buttonTest = None
memoDict = QtWidgets.QTextEdit
memoResult : QtWidgets.QTextEdit
memoUrl : QtWidgets.QTextEdit
buttonRandom = QtWidgets.QPushButton
editAntigate : QtWidgets.QLineEdit
editPause1 : QtWidgets.QLineEdit
editPause2 : QtWidgets.QLineEdit
editTime: QtWidgets.QLineEdit
buttonSaveAll : QtWidgets.QPushButton


RUCAPTCHA_TOKEN = '0ea43774452468e3d45c462b20c8e7f3'



def saveAll():
    try:
        dataSave = {
            'message': memoText.toPlainText(),
            'antigate': editAntigate.text(),
            'pause1': editPause1.text(),
            'pause2': editPause2.text(),
            'tags': memoDict.toPlainText(),
            'start': editTime.text(),
            'img': memoUrl.toPlainText(),
        }

        with open(FILE_SAVE_ALL, 'w', encoding='utf8') as outfile:
            json.dump(dataSave, outfile)
    except Exception as err:
        addLog(f'Ошибка сохранение данных, причина {err}')




def addLog(s):
    global memoLog
    memoLog.append(s)

class threadStart_(QThread):
    def critical_function(self):
        self.lock.acquire()  # заполучить лок
      #  do_something_important()  # этот код гарантированно не будет прерван каким либо другим потоком
        global memoLog
        memoLog.append('log')
          #  memoText.append('i:' + str(i))
        self.lock.release()  # освободить лок

    def run(self):
        global editPause1,editPause2, memoUrl

        opts = Options()
        opts.binary_location = "Chrome\\chrome.exe"
        dir = 'Chrome/User Data/'
        opts.add_argument("--user-data-dir=" + dir)
        opts.add_argument('--allow-profiles-outside-user-dir')
        opts.add_argument('--enable-profile-shortcut-manager')
        #opts.add_argument('--headless')

        opts.add_argument("--start-maximized")
        opts.add_experimental_option("excludeSwitches", ["enable-automation"])
        opts.add_experimental_option('useAutomationExtension', False)


        # test  - тестовый аккаунт#
        # opts.add_argument('--profile-directory=vk')  # с балансом
        #opts.add_argument('--profile-directory=test1')  # с балансом
        opts.add_argument('--profile-directory=irina')  # с балансом
        opts.add_experimental_option("useAutomationExtension", False)
        opts.add_experimental_option("excludeSwitches", ["enable-automation"])
        #browser = Chrome(executable_path="Chrome\\chromedriver_86.exe", chrome_options=opts)

        global listWidget
        global buttonStart
        global vk
        global memoText
        buttonStart.setEnabled(True)

        if listWidget.count() == 0:
            buttonStart.setEnabled(True)
            addLog('Загрузите аккаунты для рассылки')
            return 0



        for i in range(listWidget.count()):  # бежим по аккаунтам
            acc_ = ''
            try:
                acc = ''
                acc = listWidget.item(i).text()
                acc_ = listWidget.item(i).text()
                listWidget.setCurrentRow(i)




                if acc.find('com/id') != -1:
                    acc = acc[acc.find('com/id') + 6:]
                if acc.find('com/') != -1:
                    acc = acc[acc.find('com/') + 4:]
                print(acc)
                response = ''
                id = ''
                try:
                    response = vk.users.get(user_ids=acc)
                    if str(response).find('id') != -1:
                        id = str(response[0]['id'])
                        full_name = response[0]['first_name']
                        # Успешно получили ID
                except Exception as err:
                    try:
                        listWidget.item(i).setBackground(QColor('#FF9C9C'))
                        saveBad(acc_)
                        print('Ошибка отправки сообщения для ' + str(acc_) + ' Код ошибки:4 ' + str(err), err)
                        addLog('Ошибка отправки сообщения для ' + str(acc_) + ' Код ошибки:4 ' + str(err))
                        message = ''
                        continue
                    except:
                        pass
                try:
                    message = str(memoText.toPlainText())
                    text = memoText.toPlainText()
                    list = memoDict.toPlainText().split('\n')
                    message = gen_text(text,list)
                    message = str(message).replace('{username}',full_name)
                    listWidget.item(i).setBackground(QColor('#FBFF6F'))
                    # Отправка сообщения
                    #browser.get('https://vk.com/im?sel=' +id )
                    try:
                        #img
                        print('stop')

                        #отправка сообщений
                        vk.messages.send(user_id=id, random_id = random.randint(0,10000), message = message, attachment = getIMG(memoUrl.toPlainText()), captcha_handler=captcha_handler )
                        pass
                        #WebDriverWait(browser, 5).until(lambda x: x.find_element_by_css_selector('#im_editable'+str(id)))
                    except Exception as err:
                        listWidget.item(i).setBackground(QColor('#FF9C9C'))
                        saveBad(acc_)
                        print('Ошибка отправки сообщения для ' + str(acc_) + ' Код ошибки1: Не найденно поле ввода сообщения', err)
                        addLog('Ошибка отправки сообщения для ' + str(acc_) + ' Код ошибки1 Не найденно поле ввода сообщения: ' + str(err))
                        continue
                        # не нашди поле отпарвки сообщения вк

                    time.sleep(1)





                    listWidget.item(i).setBackground(QColor('#9CFFCB'))
                    addLog('Успешно отправили сообщения для ' + str(acc_))
                    # Пауза между отправкой сообщения
                    ran = random.randint(int(editPause1.text()) , int(editPause2.text()))
                    addLog('Sleep ' + str(ran) + ' sec')
                    time.sleep(ran)


                except Exception as err:
                    listWidget.item(i).setBackground(QColor('#FF9C9C'))
                    saveBad(acc_)
                    print('Ошибка отправки сообщения для ' + str(acc_) + ' Код ошибки: 3' + str(err), err)
                    addLog('Ошибка отправки сообщения для ' + str(acc_) + ' Код ошибки:3 ' + str(err))
                    continue

            except Exception as  err:
                try:
                    listWidget.item(i).setBackground(QColor('#FF9C9C'))
                    saveBad(acc_)
                    print('Ошибка отправки сообщения для ' + str(acc_) + ' Код ошибки: 3' + str(err), err)
                    addLog('Ошибка отправки сообщения для ' + str(acc_) + ' Код ошибки:3 ' + str(err))

                    continue
                except:
                    pass

        addLog('Завершили рассылку сообщений')
        # Завершили работу
        buttonStart.setEnabled(True)


def captcha_handler(captcha):
    global editAntigate
    global RUCAPTCHA_TOKEN, FILE_CAPTCHA
    addLog('Поймали капчу')
    """ При возникновении капчи вызывается эта функция и ей передается объект
        капчи. Через метод get_url можно получить ссылку на изображение.
        Через метод try_again можно попытаться отправить запрос с кодом капчи
    """
    addLog('Идёт процесс автоматического распознования капчи')
    code = ''
    url = str(captcha.get_url())
    user_answer = ImageCaptcha.ImageCaptcha(rucaptcha_key=editAntigate.text()).captcha_handler(captcha_link=url)
    if not user_answer['error']:
        # решение капчи
        print(user_answer['captchaSolve'])
        code = user_answer['captchaSolve']
        print(user_answer['taskId'])
    elif user_answer['error']:
        # Тело ошибки, если есть
        print(user_answer['errorBody'])
        print(user_answer['errorBody'])
    # Пробуем снова отправить запрос с капчей
    return captcha.try_again(code)


def threadLoadAcc():
    time.sleep(2)
    global listWidget
    listWidget.clear()
    try:
        wb = load_workbook(FILE_BASE_EXEL)
        sh = wb['Лист1']
        for row in sh.iter_rows():
            for cell in row:
                try:
                    listWidget.addItem(str(cell.hyperlink.target))
                except Exception as err:
                    print('err exla', err, cell)


    except Exception as err:
        addLog('Ошибка загрузки Аккаунтов для рассылки:1' + str( err ))

def threadAuth():
    try:
        global  vk
        vk_session = vk_api.VkApi(login = '9607526605',token = auth_vk_driver() )#captcha_handler=captcha_handler
        vk_session._auth_token()
        vk = vk_session.get_api()
        token = ''
        token = str(vk_session.token['access_token'])
        print('token', token)
        data_auth['token'] = token
        if token != '':# Успешно авторизовались
            buttonAuth.setEnabled(False)

            buttonStart.setEnabled(True)
            addLog('Успешно авторизовались')
        else: addLog('Ошибка авторизации. Пустой токен')
        print('Token vk:', data_auth)
    except Exception as err:

        print('Ошибка авторизации', err)

def startJob():
    while True:
        schedule.run_pending()
        time.sleep(1)

class MyWin(QtWidgets.QMainWindow):
    def LoadSave(self,):
        global memoText
        time.sleep(2)
        try:

            """
            dataSave = {
            'message': memoText.toPlainText(),
            'antigate': editAntigate.text(),
            'pause1': editPause1.text(),
            'pause2': editPause2.text(),
            'tags': memoDict.toPlainText(),
            'start': editTime.text(),
        }
            """

            dataLoad = {}
            with open(FILE_SAVE_ALL, 'r', encoding='utf-8') as json_file:
                dataLoad = json.load(json_file)
                pass
            memoText.setText(dataLoad['message'])
            memoDict.setText(dataLoad['tags'])
            editAntigate.setText(dataLoad['antigate'])
            editPause1.setText(dataLoad['pause1'])
            editPause2.setText(dataLoad['pause2'])
            editTime.setText(dataLoad['start'])
            memoUrl.setText(dataLoad['img'])



            pass
        except Exception as err:
            addLog(f'Ошибка открытие файла{FILE_SAVE_ALL}, причина:{err}')
        pass

    def __init__(self, parent=None):
        QtWidgets.QWidget.__init__(self, parent)
        self.ui = Ui_MainWindow()
        self.ui.setupUi(self)
        global memoLog, buttonAuth, imageCaptcha, editCaptcha, buttonCaptcha, buttonStart, buttonPause, listWidget
        global buttonLoadAccs, memoText, buttonTest
        global memoDict
        global memoResult
        global editAntigate
        global  editPause1,editPause2
        global memoUrl
        global editTime
        global buttonSaveAll


        memoUrl = self.ui.memoUrl

        editTime = self.ui.editTime

        editPause1 = self.ui.editPause1
        editPause2 = self.ui.editPause2

        editAntigate = self.ui.editantigate
        memoDict = self.ui.memoDict
        memoResult = self.ui.memoResult
        buttonRandom = self.ui.buttonRandom
        buttonRandom.clicked.connect(self.text_random)
        buttonAuth = self.ui.buttonAuth
        buttonAuth.clicked.connect(self.Auth)
        memoLog = self.ui.memoLog
        buttonStart = self.ui.buttonStart
        buttonStart.clicked.connect(self.Start)
        buttonPause = self.ui.buttonPause
        listWidget = self.ui.listWidget
        memoText = self.ui.memoText
        buttonLoadAccs = self.ui.buttonLoadAccs
        buttonLoadAccs.clicked.connect(self.loadAcc)

        buttonSaveAll = self.ui.buttonSaveAll
        buttonSaveAll.clicked.connect(saveAll)

        self.LoadSave()


    def text_random(self):
        global editAntigate,memoText,memoDict,memoResult
        try:
            text = str(memoText.toPlainText())
            list = str(memoDict.toPlainText()).split('\n')
            newText = gen_text(text,list)
            print(newText)
            memoResult.setHtml(newText)
        except Exception as err:
            print(err)


    def Start(self):
        global API_KEY,editAntigate,editPause1,editPause2,saveAll

        try:
            if len(editAntigate.text()) < 10:
                addLog('Введите Antigate')
                return

            API_KEY = str(editAntigate.text()).strip()
            print('API_KEY', API_KEY)

        except Exception as err:
            print(err)
        print('ok')

        if len( str(  editTime.text() )) == 5:

            def ss():
                self.thread = threadStart_()
                # self.thread.change_value.connect(self.setProgressVal)
                self.thread.start()

            buttonStart.setEnabled(False)
            schedule.every().day.at(editTime.text()).do(ss)
            addLog(f'Отложенный запуск в {editTime.text()}')

            pass
            return

        if len(str(editTime.text())) != 0:
            addLog('Ошибка времени, формат типа 00:00')
            return False

        self.thread = threadStart_()
        # self.thread.change_value.connect(self.setProgressVal)
        self.thread.start()
       # self.thread = QtCore.QThread()
        # создадим объект для выполнения кода в другом потоке
     #   self.browserHandler = BrowserHandler()
        # перенесём объект в другой поток
      #  self.browserHandler.moveToThread(self.thread)
        # после чего подключим все сигналы и слоты
      #  self.browserHandler.newTextAndColor.connect(self.addNewTextAndColor)
        # подключим сигнал старта потока к методу run у объекта, который должен выполнять код в другом потоке
     #   self.thread.started.connect(self.browserHandler.run)
        # запустим поток
      #  self.thread.start()


        #threadStart_(mainwindow=self).start()
       # input('next')
     #   buttonStart.setEnabled(False)
     #   threading.Thread(target=threadStart()).start()
    def loadAcc(self):
        threading.Thread(target=threadLoadAcc).start()
    def Auth(self):
        threading.Thread(target=threadAuth).start()

    threading.Thread(target=threadAuth).start()
    threading.Thread(target=threadLoadAcc).start()
    threading.Thread(target=startJob).start()


def run():
    addLog('0000000')

if __name__=="__main__":
    try:
        QtWidgets.QApplication.setAttribute(QtCore.Qt.AA_EnableHighDpiScaling, True)
        app = QtWidgets.QApplication(sys.argv)
        myapp = MyWin()
        myapp.show()
       # x = threading.Thread(target=run)
      #  x.start()
        sys.exit(app.exec_())
    except Exception as err:
        input(f'ошибка: {err}')
"""
            for i in range(0,3):
                try:
                    img = browser.find_element_by_class_name('captcha').find_element_by_css_selector('img').get_property('src')
                    file_name = 'captcha.png'
                    with open(file_name, 'wb') as file:
                        file.write(img.screenshot_as_png)

                    captcha_text = solver.solve_and_return_solution(file_name)
                    if captcha_text != 0:
                        print("Успешно решили 2 капчу: " + captcha_text)
                        addLog("Успешно решили 2 капчу: " + captcha_text)
                        good = True
                        bbbreak
                    else:
                        print("Провал( Не решили 2 капчу: " + solver.error_code + 'попытка номер' + str(i))
                        addLog("Провал( Не решили 2 капчу: " + solver.error_code + 'попытка номер' + str(i))


                except:
                    listWidget.item(i).setBackground(QColor('#FF9C9C'))
                    saveBad(acc_)
                    addLog('Ошибка разгадывания капчи')
                    break

            if not good:
                listWidget.item(i).setBackground(QColor('#FF9C9C'))
                saveBad(acc_)
                print('Ошибка разгадывания капчи ' + str(acc_) + ' Код ошибки: 3' )
                addLog('Ошибка разгадывания капчи ' + str(acc_) + ' Код ошибки:3 ')
                continue

            """